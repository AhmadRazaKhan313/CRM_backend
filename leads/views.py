import csv
import io
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from django.shortcuts import get_object_or_404
from django.http import HttpResponse

from .models import Lead, LeadActivity
from .serializers import (
    LeadListSerializer, LeadDetailSerializer,
    LeadCreateSerializer, LeadActivitySerializer
)
from core.permissions import IsManagerOrAbove, IsAnyEmployee, FeatureRequired
from notifications.utils import notify

FEATURE = FeatureRequired("leads_module")

VALID_SOURCES = ["instagram", "facebook", "linkedin", "whatsapp", "website", "email", "other"]
VALID_DEPTS   = ["sales", "tech", "seo"]
VALID_STATUS  = ["new", "contacted", "interested", "follow_up", "converted", "rejected"]

COLUMNS = [
    "full_name", "email", "phone", "contact_no",
    "country", "company",
    "source", "department", "status",
    "platform_link", "service_interest", "notes", "questionnaire",
    "instagram_url", "facebook_url", "linkedin_url",
    "lead_insta_id", "lead_fb_id", "lead_linkedin_id", "lead_whatsapp_no",
    "staff_insta_id", "staff_fb_id", "staff_linkedin_id", "staff_whatsapp_id",
]

EXPORT_HEADERS = [
    "Serial No", "ID", "Full Name", "Email", "Phone", "Contact No",
    "Country", "Company", "Source", "Department", "Status",
    "Platform Link", "Service Interest", "Notes", "Questionnaire",
    "Instagram URL", "Facebook URL", "LinkedIn URL",
    "Lead Insta ID", "Lead FB ID", "Lead LinkedIn ID", "Lead WhatsApp No",
    "Staff Insta ID", "Staff FB ID", "Staff LinkedIn ID", "Staff WhatsApp ID",
    "Assigned To", "Assigned To (Employee ID)",
    "Created By", "Created At",
]

EXAMPLE_ROW = [
    "John Doe", "john@example.com", "+1234567890", "+1234567890",
    "USA", "Acme Inc",
    "facebook", "sales", "new",
    "https://fb.com/johndoe", "Web Development", "Interested in ecommerce", "Q1: Yes",
    "", "", "https://linkedin.com/in/johndoe",
    "@lead_insta", "lead_fb_id", "lead_li_id", "+1234567890",
    "@staff_insta", "staff_fb_id", "staff_li_id", "+0987654321",
]

LEAD_PATCH_ALLOWED = {
    "full_name", "email", "phone", "contact_no", "country", "company",
    "source", "department", "status", "service_interest", "notes",
    "questionnaire", "platform_link", "serial_no",
    "instagram_url", "facebook_url", "linkedin_url",
    "lead_insta_id", "lead_fb_id", "lead_linkedin_id", "lead_whatsapp_no",
    "staff_insta_id", "staff_fb_id", "staff_linkedin_id", "staff_whatsapp_id",
    "assigned_to",
}


# ✅ Helper — Lead se Client auto-create karo
def _auto_create_client(lead, user):
    """
    When lead status becomes 'converted' — client is created automatically.
    Skips if client already exists.
    """
    from clients.models import Client

    # Skip if client already exists
    if lead.converted_client.exists():
        return lead.converted_client.first()

    client = Client.objects.create(
        tenant         = lead.tenant,
        created_by     = user,
        assigned_to    = lead.assigned_to,
        converted_from = lead,
        full_name      = lead.full_name,
        email          = lead.email,
        phone          = lead.phone or lead.contact_no,
        country        = lead.country,
        company        = lead.company,
        department     = lead.department,
        notes          = lead.notes,
        status         = "active",
    )

    LeadActivity.objects.create(
        lead          = lead,
        activity_type = "status_change",
        note          = f"Lead converted — client created automatically (ID: {client.id})",
        created_by    = user,
    )

    return client


def parse_row(row, index, tenant, user):
    errors = []
    full_name = str(row.get("full_name", "") or "").strip()
    source    = str(row.get("source",    "") or "").strip().lower()
    dept      = str(row.get("department","") or "").strip().lower()
    status_v  = str(row.get("status",   "new") or "new").strip().lower() or "new"

    if not full_name:               errors.append("full_name is empty")
    if source not in VALID_SOURCES: errors.append(f"invalid source '{source}'")
    if dept not in VALID_DEPTS:     errors.append(f"invalid department '{dept}'")
    if status_v not in VALID_STATUS: status_v = "new"

    if errors:
        return None, errors

    data = dict(
        tenant            = tenant,
        created_by        = user,
        full_name         = full_name,
        email             = str(row.get("email",            "") or "").strip(),
        phone             = str(row.get("phone",            "") or "").strip(),
        contact_no        = str(row.get("contact_no",       "") or "").strip(),
        country           = str(row.get("country",          "") or "").strip(),
        company           = str(row.get("company",          "") or "").strip(),
        source            = source,
        department        = dept,
        status            = status_v,
        platform_link     = str(row.get("platform_link",    "") or "").strip(),
        service_interest  = str(row.get("service_interest", "") or "").strip(),
        notes             = str(row.get("notes",            "") or "").strip(),
        questionnaire     = str(row.get("questionnaire",    "") or "").strip(),
        instagram_url     = str(row.get("instagram_url",    "") or "").strip(),
        facebook_url      = str(row.get("facebook_url",     "") or "").strip(),
        linkedin_url      = str(row.get("linkedin_url",     "") or "").strip(),
        lead_insta_id     = str(row.get("lead_insta_id",    "") or "").strip(),
        lead_fb_id        = str(row.get("lead_fb_id",       "") or "").strip(),
        lead_linkedin_id  = str(row.get("lead_linkedin_id", "") or "").strip(),
        lead_whatsapp_no  = str(row.get("lead_whatsapp_no", "") or "").strip(),
        staff_insta_id    = str(row.get("staff_insta_id",   "") or "").strip(),
        staff_fb_id       = str(row.get("staff_fb_id",      "") or "").strip(),
        staff_linkedin_id = str(row.get("staff_linkedin_id","") or "").strip(),
        staff_whatsapp_id = str(row.get("staff_whatsapp_id","") or "").strip(),
    )
    return data, []


def lead_to_row(l):
    return [
        l.serial_no, l.id, l.full_name, l.email, l.phone, l.contact_no,
        l.country, l.company, l.source, l.department, l.status,
        l.platform_link, l.service_interest, l.notes, l.questionnaire,
        l.instagram_url, l.facebook_url, l.linkedin_url,
        l.lead_insta_id, l.lead_fb_id, l.lead_linkedin_id, l.lead_whatsapp_no,
        l.staff_insta_id, l.staff_fb_id, l.staff_linkedin_id, l.staff_whatsapp_id,
        l.assigned_to.full_name if l.assigned_to else "",
        getattr(l.assigned_to, "employee_id", "") if l.assigned_to else "",
        l.created_by.full_name if l.created_by else "",
        l.created_at.strftime("%Y-%m-%d %H:%M"),
    ]


def get_filtered_qs(request):
    qs = Lead.objects.filter(
        tenant=request.user.tenant,
        is_archived=False
    ).select_related("assigned_to", "created_by")

    if request.query_params.get("status"):
        qs = qs.filter(status=request.query_params["status"])
    if request.query_params.get("department"):
        qs = qs.filter(department=request.query_params["department"])
    if request.query_params.get("source"):
        qs = qs.filter(source=request.query_params["source"])
    if request.user.role in ("dept_head", "lead_manager", "sales_manager"):
        qs = qs.filter(department=request.user.department)
    return qs


class LeadListCreateView(APIView):
    permission_classes = (IsAnyEmployee, FEATURE)

    def get(self, request):
        qs = Lead.objects.filter(
            tenant=request.user.tenant,
            is_archived=False
        ).select_related("assigned_to", "created_by")

        if request.query_params.get("status"):     qs = qs.filter(status=request.query_params["status"])
        if request.query_params.get("department"): qs = qs.filter(department=request.query_params["department"])
        if request.query_params.get("source"):     qs = qs.filter(source=request.query_params["source"])
        if request.query_params.get("search"):
            s = request.query_params["search"]
            qs = qs.filter(full_name__icontains=s) | qs.filter(country__icontains=s)

        if request.user.is_super_admin or request.user.role in ("ceo", "coo", "sales_director"):
            return Response(LeadListSerializer(qs, many=True).data)
        if request.user.role in ("dept_head", "lead_manager", "sales_manager"):
            qs = qs.filter(department=request.user.department)
        elif request.user.role in ("lead_employee", "sales_employee"):
            qs = qs.filter(assigned_to=request.user)

        return Response(LeadListSerializer(qs, many=True).data)

    def post(self, request):
        serializer = LeadCreateSerializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        lead = serializer.save()

        # ✅ Agar lead directly "converted" status ke saath bani toh client auto-create
        if lead.status == "converted":
            _auto_create_client(lead, request.user)

        return Response(LeadDetailSerializer(lead).data, status=status.HTTP_201_CREATED)


class LeadDetailView(APIView):
    permission_classes = (IsAnyEmployee, FEATURE)

    def _get_lead(self, pk, user):
        lead = get_object_or_404(Lead, pk=pk, tenant=user.tenant, is_archived=False)
        if user.role in ("lead_employee", "sales_employee"):
            if lead.assigned_to != user:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied()
        return lead

    def get(self, request, pk):
        return Response(LeadDetailSerializer(self._get_lead(pk, request.user)).data)

    def patch(self, request, pk):
        lead       = self._get_lead(pk, request.user)
        old_status = lead.status
        new_status = request.data.get("status")

        # ✅ Converted lead ka status wapas change nahi ho sakta
        if old_status == "converted" and new_status and new_status != "converted":
            if lead.converted_client.exists():
                return Response(
                    {"detail": "This lead has already been converted — status cannot be changed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        safe_data  = {k: v for k, v in request.data.items() if k in LEAD_PATCH_ALLOWED}
        serializer = LeadCreateSerializer(
            lead, data=safe_data, partial=True, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        lead.refresh_from_db()

        # ✅ Status "converted" ho gaya — client auto-create
        if new_status == "converted" and old_status != "converted":
            _auto_create_client(lead, request.user)

        # Activity log
        if new_status and new_status != old_status:
            LeadActivity.objects.create(
                lead          = lead,
                activity_type = "status_change",
                note          = f"Status changed from {old_status} to {lead.status}",
                created_by    = request.user
            )

        return Response(LeadDetailSerializer(lead).data)

    def delete(self, request, pk):
        lead = self._get_lead(pk, request.user)
        lead.is_archived = True
        lead.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class LeadAssignView(APIView):
    permission_classes = (IsManagerOrAbove,)

    def post(self, request, pk):
        lead = get_object_or_404(Lead, pk=pk, tenant=request.user.tenant)
        from authentication.models import User
        employee = get_object_or_404(User, pk=request.data.get("user_id"), tenant=request.user.tenant)
        lead.assigned_to = employee
        lead.save()
        LeadActivity.objects.create(
            lead          = lead,
            activity_type = "note",
            note          = f"Lead assigned to {employee.full_name}",
            created_by    = request.user
        )
        return Response(LeadDetailSerializer(lead).data)


class LeadActivityView(APIView):
    permission_classes = (IsAnyEmployee, FEATURE)

    def post(self, request, pk):
        lead     = get_object_or_404(Lead, pk=pk, tenant=request.user.tenant)
        activity = LeadActivity.objects.create(
            lead          = lead,
            activity_type = request.data.get("activity_type", "note"),
            note          = request.data.get("note", ""),
            created_by    = request.user
        )
        return Response(LeadActivitySerializer(activity).data, status=status.HTTP_201_CREATED)


class LeadConvertView(APIView):
    """
    POST /api/leads/<pk>/convert/
    Can also be used for manual convert button.
    """
    permission_classes = (IsManagerOrAbove, FEATURE)

    def post(self, request, pk):
        lead = get_object_or_404(Lead, pk=pk, tenant=request.user.tenant, is_archived=False)

        existing = lead.converted_client.first()
        if existing:
            from clients.serializers import ClientDetailSerializer
            return Response({
                "detail": "Lead has already been converted.",
                "client_id": existing.id,
            }, status=status.HTTP_200_OK)

        lead.status = "converted"
        lead.save()

        client = _auto_create_client(lead, request.user)
        from clients.serializers import ClientDetailSerializer

        return Response({
            "detail": "Lead successfully converted.",
            "client_id": client.id,
            "client": ClientDetailSerializer(client).data,
        }, status=status.HTTP_201_CREATED)


class LeadBulkUploadView(APIView):
    permission_classes = (IsManagerOrAbove, FEATURE)
    parser_classes     = (MultiPartParser,)

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "File required."}, status=400)

        fname = file.name.lower()

        if fname.endswith(".csv"):
            rows, err = self._parse_csv(file)
        elif fname.endswith((".xlsx", ".xls")):
            rows, err = self._parse_excel(file)
        else:
            return Response({"detail": "Only .csv, .xlsx, .xls files allowed."}, status=400)

        if err:
            return Response({"detail": err}, status=400)

        created, errors = [], []

        for i, row in enumerate(rows, start=2):
            data, row_errors = parse_row(row, i, request.user.tenant, request.user)
            if row_errors:
                errors.append({"row": i, "name": str(row.get("full_name", f"Row {i}")), "errors": row_errors})
                continue
            try:
                Lead.objects.create(**data)
                created.append(i)
            except Exception as e:
                errors.append({"row": i, "name": data.get("full_name", ""), "errors": [str(e)]})

        return Response({
            "created": len(created),
            "failed":  len(errors),
            "errors":  errors[:20],
        }, status=201 if created else 400)

    def _parse_csv(self, file):
        try:
            decoded = file.read().decode("utf-8-sig")
        except UnicodeDecodeError:
            return [], "File encoding error. Save as UTF-8 CSV."
        reader  = csv.DictReader(io.StringIO(decoded))
        headers = reader.fieldnames or []
        missing = [f for f in ["full_name", "source", "department"] if f not in headers]
        if missing:
            return [], f"Missing columns: {', '.join(missing)}"
        return list(reader), None

    def _parse_excel(self, file):
        try:
            import openpyxl
            wb        = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws        = wb.active
            rows_iter = iter(ws.rows)
            headers   = [str(cell.value or "").strip() for cell in next(rows_iter)]
            missing   = [f for f in ["full_name", "source", "department"] if f not in headers]
            if missing:
                return [], f"Missing columns: {', '.join(missing)}"
            rows = []
            for excel_row in rows_iter:
                row_dict = {headers[i]: (str(cell.value) if cell.value is not None else "") for i, cell in enumerate(excel_row) if i < len(headers)}
                rows.append(row_dict)
            return rows, None
        except ImportError:
            return [], "openpyxl not installed."
        except Exception as e:
            return [], f"Excel parse error: {str(e)}"


class LeadTemplateDownloadView(APIView):
    permission_classes = (IsManagerOrAbove, FEATURE)

    def get(self, request):
        fmt = request.query_params.get("format", "csv").lower()
        if fmt == "excel":
            return self._excel_template()
        return self._csv_template()

    def _csv_template(self):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="leads_template.csv"'
        w = csv.writer(response)
        w.writerow(COLUMNS)
        w.writerow(EXAMPLE_ROW)
        return response

    def _excel_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=500)

        wb    = openpyxl.Workbook()
        ws    = wb.active
        ws.title = "Leads Template"
        hfill = PatternFill("solid", fgColor="4F6EF7")
        hfont = Font(bold=True, color="FFFFFF")

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 15)

        for col_idx, val in enumerate(EXAMPLE_ROW, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        ws.cell(row=3, column=1, value="↑ Delete example rows before uploading")
        ws.cell(row=3, column=1).font = Font(italic=True, color="999999")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="leads_template.xlsx"'
        return response


class LeadExportView(APIView):
    permission_classes = (IsManagerOrAbove, FEATURE)

    def get(self, request):
        fmt = request.query_params.get("format", "csv").lower()
        qs  = get_filtered_qs(request)
        if fmt == "excel":
            return self._export_excel(qs)
        return self._export_csv(qs)

    def _export_csv(self, qs):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="leads_export.csv"'
        w = csv.writer(response)
        w.writerow(EXPORT_HEADERS)
        for l in qs:
            w.writerow(lead_to_row(l))
        return response

    def _export_excel(self, qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse("openpyxl not installed.", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title       = "Leads Export"
        ws.freeze_panes = "A2"
        hfill = PatternFill("solid", fgColor="4F6EF7")
        hfont = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = hfill
            cell.font      = hfont
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)

        status_colors = {
            "new":        "DBEAFE",
            "contacted":  "FEF9C3",
            "interested": "DCFCE7",
            "follow_up":  "FFEDD5",
            "converted":  "EDE9FE",
            "rejected":   "FEE2E2",
        }

        for row_idx, lead in enumerate(qs, start=2):
            row_data  = lead_to_row(lead)
            row_color = status_colors.get(lead.status, "FFFFFF")
            row_fill  = PatternFill("solid", fgColor=row_color)
            for col_idx, value in enumerate(row_data, start=1):
                cell      = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="leads_export.xlsx"'
        return response